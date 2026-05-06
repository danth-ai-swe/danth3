"""Parse Knowledge-Node and Quiz xlsx files, plus syllabus sheets."""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from loma_rag.chunking.text import norm_ws
from loma_rag.ingest.filename import parse_filename
from loma_rag.model.domain import IngestNode, Quiz


def parse_nodes(path: Path) -> list[IngestNode]:
    meta = parse_filename(path.name)
    if not meta:
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c else "" for c in rows[0]]

    def col(name: str) -> int:
        for i, h in enumerate(header):
            if h.lower() == name.lower():
                return i
        return -1

    i_id = col("Node ID")
    i_name = col("Node Name")
    i_def = col("Definition")
    i_cat = col("Category")
    i_tags = col("Domain Tags")
    i_rel = col("Related Nodes")

    if i_id < 0 or i_name < 0:
        return []

    def cell(row: tuple, idx: int) -> str:
        if idx < 0 or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    out: list[IngestNode] = []
    for row in rows[1:]:
        if not row or row[i_id] in (None, ""):
            continue
        raw_id = cell(row, i_id)
        out.append(IngestNode(
            node_id=f"{meta['lesson_id']}::{raw_id}",
            raw_id=raw_id,
            name=cell(row, i_name),
            definition=cell(row, i_def),
            category=cell(row, i_cat),
            domain_tags=cell(row, i_tags),
            related_raw=cell(row, i_rel),
            course=meta["course"],
            module=meta["module"],
            lesson=meta["lesson"],
            lesson_id=meta["lesson_id"],
            source=path.name,
        ))
    return out


def parse_quiz(path: Path) -> list[Quiz]:
    course_folder = path.parents[1].name  # <course>/03_Quiz/<file>.xlsx
    if "281" in course_folder:
        course = "LOMA281"
    elif "291" in course_folder:
        course = "LOMA291"
    else:
        course = "UNKNOWN"
    m = re.search(r"Module\s*(\d+)\s*Lesson\s*(\d+)", path.stem, re.I)
    if not m:
        return []
    module = f"M{m.group(1)}"
    lesson = f"L{m.group(2)}"
    lesson_id = f"{course}_{module}{lesson}"

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []

    # Detect column layout from the header row. Most files have 8 cols
    # (correct_col=6), but a few (e.g. M3L1) have 9 cols with a 5th option slot.
    header = rows[0]

    def _find_col(name: str) -> int:
        for i, c in enumerate(header):
            if c is not None and str(c).strip().lower() == name.lower():
                return i
        return -1

    correct_col = _find_col("Correct Answer")
    diff_col = _find_col("Difficulty Level")
    if correct_col < 0:
        correct_col = 6  # legacy fallback
    if diff_col < 0:
        diff_col = correct_col + 1

    out: list[Quiz] = []
    # Skip first 2 header rows.
    for r in rows[2:]:
        if not r or r[0] is None:
            continue
        try:
            no = int(r[0])
        except (ValueError, TypeError):
            continue
        question = str(r[1]).strip() if r[1] else ""
        options = []
        for k in range(2, correct_col):
            v = r[k] if k < len(r) else None
            options.append(str(v).strip() if v is not None and str(v).strip() else "")
        try:
            cv = r[correct_col] if correct_col < len(r) else None
            correct_idx = int(cv) if cv is not None else 0
        except (ValueError, TypeError):
            correct_idx = 0
        correct_text = (
            options[correct_idx - 1]
            if 1 <= correct_idx <= len(options)
            else ""
        )
        difficulty = (
            str(r[diff_col]).strip()
            if diff_col >= 0 and diff_col < len(r) and r[diff_col]
            else ""
        )
        out.append(Quiz(
            quiz_id=f"{lesson_id}_Q{no}",
            course=course,
            module=module,
            lesson=lesson,
            lesson_id=lesson_id,
            no=no,
            question=question,
            options=options,
            correct_idx=correct_idx,
            correct_text=correct_text,
            difficulty=difficulty,
            source=path.name,
        ))
    return out


def _parse_syllabus_meta(ws) -> dict:
    """Pull (Key Name -> Value) from the 'Syllabus' sheet's row layout."""
    out: dict = {}
    for r in ws.iter_rows(values_only=True):
        if len(r) < 4:
            continue
        key = str(r[2]).strip() if r[2] else ""
        val = r[3]
        if key and val is not None:
            clean_key = key.rstrip().rstrip("*").rstrip()
            out[clean_key] = str(val).strip()
    return out


def _parse_schedule(ws) -> dict:
    """Walk the 'Schedule' sheet into structured modules/lessons/totals."""
    import re as _re
    _M_RE = _re.compile(r"^Module\s+(\d+):\s*(.+)", _re.S)
    _L_RE = _re.compile(r"^Lesson\s+(\d+):\s*(.+)", _re.S)

    rows = list(ws.iter_rows(values_only=True))
    modules: list = []
    current_module: dict | None = None
    totals: dict = {}

    for r in rows:
        c1 = str(r[1]).strip() if len(r) > 1 and r[1] else ""
        c2 = str(r[2]).strip() if len(r) > 2 and r[2] else ""
        c3 = str(r[3]).strip() if len(r) > 3 and r[3] else ""
        c4 = str(r[4]).strip() if len(r) > 4 and r[4] else ""
        c5 = r[5] if len(r) > 5 else None
        c9 = str(r[9]).strip() if len(r) > 9 and r[9] else ""

        # Module header (col 1).
        if c1.startswith("Module "):
            mm = _M_RE.match(c1)
            if mm:
                full = mm.group(2).strip()
                title_line = full.split("\n")[0].strip()
                desc = full[len(title_line):].strip()
                current_module = {
                    "num": int(mm.group(1)),
                    "name": title_line,
                    "description": desc,
                    "lessons": [],
                }
                modules.append(current_module)

        # Lesson row.
        if c2.startswith("Lesson ") and current_module is not None:
            lm = _L_RE.match(c2)
            if lm:
                sections = [s.strip() for s in c3.split("\n") if s.strip()] if c3 else []
                try:
                    duration = float(c5) if c5 is not None else 0.0
                except (TypeError, ValueError):
                    duration = 0.0
                current_module["lessons"].append({
                    "num": int(lm.group(1)),
                    "name": lm.group(2).strip(),
                    "sections": sections,
                    "delivery_mode": c4,
                    "duration": duration,
                    "detail": c9,
                })

        # Totals at bottom (col 1 has the label, col 2 has the value).
        if c1 in ("Document self-learning", "Quiz", "Review", "Total hours") and not c2.startswith("Lesson "):
            try:
                totals[c1] = float(r[2]) if r[2] is not None else None
            except (TypeError, ValueError):
                pass

    return {"modules": modules, "totals": totals}
