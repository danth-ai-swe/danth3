"""Scan course/module/lesson directory trees to build TOC + syllabus chunks."""
from __future__ import annotations

import re
from pathlib import Path

from docx import Document
from openpyxl import load_workbook

from loma_rag.chunking.text import norm_ws
from loma_rag.ingest.xlsx_parser import _parse_syllabus_meta, _parse_schedule
from loma_rag.model.domain import Chunk

_COURSE_DIR_RE = re.compile(r"^LOMA\s+(\d+)\s+-\s+(.+)$")
_MODULE_DIR_RE = re.compile(r"^Module\s+(\d+)\s+(.+)$")
_LESSON_DIR_RE = re.compile(r"^Lesson\s+(\d+)\s*-\s*(.+)$")

_M_RE = re.compile(r"^Module\s+(\d+):\s*(.+)", re.S)
_L_RE = re.compile(r"^Lesson\s+(\d+):\s*(.+)", re.S)


def _extract_learning_objectives(lesson_dir: Path) -> list[str]:
    """Open the Knowledge File docx and return the list of bullet points that
    appear under the 'Learning Objectives' Heading 3. Empty list on failure."""
    docx_files = [p for p in lesson_dir.glob("*.docx") if "Knowledge File" in norm_ws(p.name)]
    if not docx_files:
        return []
    try:
        d = Document(docx_files[0])
    except Exception:
        return []
    bullets: list[str] = []
    in_lo = False
    for p in d.paragraphs:
        style = (p.style.name or "") if p.style else ""
        text = p.text.strip()
        if not text:
            continue
        if style == "Heading 3":
            # Enter LO section if heading text matches; exit on next H3.
            in_lo = "learning objective" in text.lower()
            continue
        if style.startswith("Heading"):
            in_lo = False
            continue
        if in_lo:
            # Skip the lead-in line.
            if text.lower().startswith("after studying this lesson"):
                continue
            bullets.append(text)
            if len(bullets) >= 6:  # cap per lesson to keep chunks compact
                break
    return bullets


def _scan_course_tree(root: Path) -> dict:
    """Returns: {course_id: {"full_name":..., "modules": {n: {"name":..., "lessons":[{num,name,objectives}]}}}}"""
    out: dict = {}
    for c_dir in sorted(root.iterdir()):
        if not c_dir.is_dir():
            continue
        cm = _COURSE_DIR_RE.match(c_dir.name)
        if not cm:
            continue
        course_id = f"LOMA{cm.group(1)}"
        modules: dict = {}
        doc_dir = c_dir / "02_Document"
        if not doc_dir.exists():
            continue
        for m_dir in sorted(doc_dir.iterdir()):
            if not m_dir.is_dir():
                continue
            mm = _MODULE_DIR_RE.match(m_dir.name)
            if not mm:
                continue
            m_num = int(mm.group(1))
            lessons = []
            for l_dir in sorted(m_dir.iterdir()):
                if not l_dir.is_dir():
                    continue
                lm = _LESSON_DIR_RE.match(l_dir.name)
                if lm:
                    lessons.append({
                        "num": int(lm.group(1)),
                        "name": lm.group(2).strip(),
                        "objectives": _extract_learning_objectives(l_dir),
                    })
            modules[m_num] = {"name": mm.group(2).strip(), "lessons": lessons}
        out[course_id] = {"full_name": c_dir.name, "modules": modules}
    return out


def build_toc_chunks(root: Path) -> list[Chunk]:
    """Generate retrievable 'table of contents' chunks describing the course
    structure (#modules, #lessons, lesson titles per module). Includes both
    English and Vietnamese phrasings so structural questions retrieve well."""
    chunks: list[Chunk] = []
    courses = _scan_course_tree(root)

    for course_id, info in courses.items():
        modules = info["modules"]
        n_modules = len(modules)
        n_lessons = sum(len(m["lessons"]) for m in modules.values())

        # Course-level TOC.
        lines = [
            f"# {course_id} — {info['full_name']}",
            "",
            f"This course is organised into {n_modules} modules with a total of "
            f"{n_lessons} lessons.",
            f"Khóa học {course_id} gồm {n_modules} modules với tổng cộng "
            f"{n_lessons} lessons (bài học).",
            "",
            "## Course outline / Mục lục khóa học:",
        ]
        for m_num in sorted(modules):
            m = modules[m_num]
            lines.append(
                f"\nModule {m_num}: {m['name']} "
                f"(contains {len(m['lessons'])} lessons / có {len(m['lessons'])} bài)"
            )
            for ld in m["lessons"]:
                lines.append(f"  - Lesson {ld['num']}: {ld['name']}")
        text = "\n".join(lines)

        chunks.append(Chunk(
            chunk_id=f"{course_id}_TOC",
            course=course_id,
            module="*",
            lesson="*",
            lesson_id=f"{course_id}_TOC",
            section="Course Outline",
            subsection="",
            text=text,
            char_count=len(text),
            token_estimate=len(text) // 4,
            source=f"{course_id}_directory_scan",
        ))

        # Per-module TOC chunk — includes each lesson's learning objectives so
        # questions about lesson CONTENT can be answered from this single chunk.
        for m_num in sorted(modules):
            m = modules[m_num]
            lines = [
                f"# {course_id} — Module {m_num}: {m['name']}",
                "",
                f"Module {m_num} of course {course_id} ({info['full_name']}).",
                f"Module {m_num} của khóa {course_id} có tựa đề là "
                f'"{m["name"]}" và bao gồm {len(m["lessons"])} lessons sau.',
                "",
            ]
            for ld in m["lessons"]:
                lines.append(f"## Lesson {ld['num']}: {ld['name']}")
                if ld["objectives"]:
                    lines.append("Learning objectives / Mục tiêu học tập:")
                    for o in ld["objectives"]:
                        lines.append(f"- {o}")
                lines.append("")
            lines.append(
                f"In total, module {m_num} contains {len(m['lessons'])} lessons. "
                f"Tổng cộng module {m_num} có {len(m['lessons'])} bài."
            )
            text = "\n".join(lines)
            chunks.append(Chunk(
                chunk_id=f"{course_id}_M{m_num}_TOC",
                course=course_id,
                module=f"M{m_num}",
                lesson="*",
                lesson_id=f"{course_id}_M{m_num}_TOC",
                section=f"Module {m_num} Outline",
                subsection=m["name"],
                text=text,
                char_count=len(text),
                token_estimate=len(text) // 4,
                source=f"{course_id}_directory_scan",
            ))
    return chunks


def _find_syllabus(course_dir: Path) -> Path | None:
    syl_dir = course_dir / "01_Syllabus"
    if not syl_dir.is_dir():
        return None
    for p in sorted(syl_dir.glob("*.xlsx")):
        return p
    return None


def build_syllabus_chunks(root: Path) -> list[Chunk]:
    """Parse 01_Syllabus/*.xlsx for each course → produce:
       - 1 course-level metadata chunk (Topic, Method, Objectives, Outcomes, totals)
       - 1 per-module schedule chunk (lesson sections, durations, delivery mode)
    Both are highly retrievable for course-meta questions ("how many hours",
    "training method", "what sections in lesson X")."""
    chunks: list[Chunk] = []
    courses = _scan_course_tree(root)

    for course_id, info in courses.items():
        c_dir = root / info["full_name"]
        syllabus_path = _find_syllabus(c_dir)
        if not syllabus_path:
            continue
        try:
            wb = load_workbook(syllabus_path, data_only=True)
        except Exception:
            continue

        meta: dict = {}
        sched: dict = {"modules": [], "totals": {}}
        if "Syllabus" in wb.sheetnames:
            meta = _parse_syllabus_meta(wb["Syllabus"])
        if "Schedule" in wb.sheetnames:
            sched = _parse_schedule(wb["Schedule"])

        # Course-level syllabus chunk. Skips "Course Outlines" (covered by
        # TOC chunks) to avoid lexical overlap that would let TOC chunks
        # crowd out syllabus chunks for course-meta questions.
        total_hours = sched["totals"].get("Total hours")
        method = meta.get("Training Method(s)", "")
        learners = meta.get("Type of Learners", "")
        passing = meta.get("Passing criteria", "")

        lines = [
            f"# {course_id} — Course Syllabus / Đề cương khóa học {course_id}",
            f"Source: {syllabus_path.name}",
            "",
            (
                f"This chunk answers course-level meta questions about {course_id}: "
                f"course duration / thời lượng khóa học / tổng số giờ / how many hours, "
                f"training method / phương pháp học / cách học, "
                f"course objectives / mục tiêu khóa học, "
                f"course outcomes / kết quả đầu ra / outcome, "
                f"prerequisites / yêu cầu tiên quyết / kiến thức cần có, "
                f"type of learners / đối tượng học viên, "
                f"passing criteria / tiêu chí đậu / điều kiện hoàn thành / qua môn, "
                f"assessment / đánh giá / chấm điểm / hình thức kiểm tra."
            ),
            "",
            f"## Course metadata / Thông tin chung khóa học {course_id}",
        ]
        if total_hours is not None:
            lines.append(
                f"- Total duration / Tổng thời lượng / Tổng số giờ: "
                f"{total_hours} hours ({total_hours} giờ)."
            )
        if method:
            lines.append(
                f"- Training method / Phương pháp đào tạo / Cách học: {method}."
            )
        if learners:
            lines.append(
                f"- Type of learners / Đối tượng học viên: {learners}."
            )
        if "Prerequisites of knowledge" in meta:
            lines.append(
                f"- Prerequisites / Yêu cầu tiên quyết / Kiến thức yêu cầu: "
                f"{meta['Prerequisites of knowledge']}."
            )
        if passing:
            lines.append(
                f"- Passing criteria / Tiêu chí đậu / Điều kiện hoàn thành: {passing}."
            )
        if "Assessment Scheme" in meta:
            lines.append(
                f"- Assessment / Hình thức đánh giá: {meta['Assessment Scheme']}."
            )
        lines.append("")

        if "Course Objectives" in meta:
            lines.append("## Course Objectives / Mục tiêu khóa học")
            lines.append(meta["Course Objectives"])
            lines.append("")
        if "Course Outcomes" in meta:
            lines.append("## Course Outcomes / Kết quả đầu ra / Outcomes")
            lines.append(meta["Course Outcomes"])
            lines.append("")

        if sched["totals"]:
            lines.append("## Schedule totals / Thời lượng theo loại hoạt động")
            for k, v in sched["totals"].items():
                if v is not None:
                    lines.append(f"- {k}: {v} hours / {v} giờ")
            lines.append("")

        text = "\n".join(lines).strip()
        if text:
            chunks.append(Chunk(
                chunk_id=f"{course_id}_SYLLABUS",
                course=course_id,
                module="*",
                lesson="*",
                lesson_id=f"{course_id}_SYLLABUS",
                section="Course Syllabus",
                subsection="",
                text=text,
                char_count=len(text),
                token_estimate=len(text) // 4,
                source=syllabus_path.name,
            ))

        # Per-module schedule chunk.
        for m in sched["modules"]:
            m_num = m["num"]
            m_name = m["name"]
            module_total = sum(ld["duration"] for ld in m["lessons"])
            lines = [
                f"# {course_id} — Module {m_num} Schedule: {m_name}",
                f"Source: {syllabus_path.name}",
                "",
                "## Module metadata / Lịch học module",
                f"- Number of lessons / Số bài học: {len(m['lessons'])}",
                f"- Self-study total / Tổng thời lượng module / Module {m_num} mất bao lâu: "
                f"~{module_total} hours / {module_total} giờ (chưa kể quiz và review).",
                "",
            ]
            if m["description"]:
                lines.append("## Module description / Giới thiệu module")
                lines.append(m["description"])
                lines.append("")
            for ld in m["lessons"]:
                lines.append(f"## Lesson {ld['num']}: {ld['name']}")
                if ld["sections"]:
                    lines.append("Sections covered / Các phần / Mục tiêu của lesson:")
                    for s in ld["sections"]:
                        lines.append(f"- {s}")
                if ld["delivery_mode"]:
                    lines.append(
                        f"Delivery mode / Phương pháp / Cách học: {ld['delivery_mode']}."
                    )
                if ld["duration"]:
                    lines.append(
                        f"Duration / Thời lượng / Số giờ: {ld['duration']} hours "
                        f"({ld['duration']} giờ)."
                    )
                if ld["detail"]:
                    lines.append(f"Materials / Tài liệu: {ld['detail']}")
                lines.append("")
            text = "\n".join(lines).strip()
            chunks.append(Chunk(
                chunk_id=f"{course_id}_M{m_num}_SCHEDULE",
                course=course_id,
                module=f"M{m_num}",
                lesson="*",
                lesson_id=f"{course_id}_M{m_num}_SCHEDULE",
                section=f"Module {m_num} Schedule",
                subsection=m_name,
                text=text,
                char_count=len(text),
                token_estimate=len(text) // 4,
                source=syllabus_path.name,
            ))
    return chunks
